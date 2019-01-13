import sys
import os
import time
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from kingWeb.util.SqlHelper import *
from kingWeb.models import *
from django.db.models import Q
from kingWeb.settings import ROOT_PATH

class SysHelper(object):
        userid = '' #记录当前登陆用户id


        @staticmethod
        def get_column_names(tableid,condition,orderby,returnlist = True):
            """     returnlist  是否返回数组，为false返回逗号隔开的字符串
            """
            sql = 'select * from Sys_TableColumn where TableId={0} and {1} order By {2}'
            column_data = SqlHelper.query(sql.format(tableid,condition,orderby))
            column_names = []
            for row in column_data:
                column_names.append(row['Name'])
            if returnlist:
                return column_names
            else:
                return ','.join(column_names)

        @staticmethod
        def get_out_list(outsql):
            outdata_arr = outsql.split('|') #Example: Id,Name|Sys_Department|ParentId=0
            colnames = outdata_arr[0].split(',') # value,text
            tablename = outdata_arr[1]
            condition = outdata_arr[2].replace('{UserId}',SysHelper.userid)
            primarkey = colnames[0] #作为下拉菜单value的列
            textkey = colnames[1] #作为下拉菜单的text的列
            outdatalist = SqlHelper.query('select {0} as value,{1} as text from {2} where {3}'.\
                format(primarkey,textkey,tablename,condition))
            return outdatalist

        #获取外键id值对应的值
        @staticmethod
        def get_out_value(tableid,outcolname,outvalueid):
            if outvalueid == '0' or outvalueid == 0:
                return '无'
            sql = "select OutSql from Sys_TableColumn where TableId={0} and Name='{1}'"
            outdata = SqlHelper.single(sql.format(tableid,outcolname))
            outdata_arr = outdata.split('|') #Example: Id,Name|Sys_Department|ParentId=0
            colnames = outdata_arr[0].split(',') # value,text
            tablename = outdata_arr[1]
            condition = outdata_arr[2]
            primarkey = colnames[0] #作为下拉菜单value的列
            textkey = colnames[1] #作为下拉菜单的text的列
            value = SqlHelper.single('select {0} from {1} where {2}={3}'.format(textkey,tablename,primarkey,outvalueid))
            if not value:
                return '无'
            return value

        #获取外键值对应的外键id
        @staticmethod
        def get_out_value_id(tableid,outcolname,colvalue):
            if not colvalue or colvalue == 0:
                return '无'
            sql = "select OutSql from Sys_TableColumn where TableId={0} and Name='{1}'"
            outdata = SqlHelper.single(sql.format(tableid,outcolname))
            outdata_arr = outdata.split('|') #Example: Id,Name|Sys_Department|ParentId=0
            colnames = outdata_arr[0].split(',') # value,text
            tablename = outdata_arr[1]
            condition = outdata_arr[2]
            primarkey = colnames[0] #作为下拉菜单value的列
            textkey = colnames[1] #作为下拉菜单的text的列
            value = SqlHelper.single("select {0} from {1} where {2}='{3}'".format(primarkey,tablename,textkey,colvalue))
            if not value:
                return '无'
            return value
        #获取checkbox/radio选中列表中 这个值value对应的text
        #valueids 为逗号隔开的多个值
        @staticmethod
        def get_select_value(tableid,colname,valueids):
            if valueids == 0 or not valueids:
                return '无'
            column = SysTableColumn.objects.filter(Q(tableid=int(tableid)) & Q(name=colname)).first()
            option_data = column.selectrange.split('|') #value,text|value,text
            values = valueids.split(',')
            result = ''
            for option in option_data:
                val = option.split(',')[0]
                if val in values:
                    result+=option.split(',')[1] + ','
            return result.strip(',')
        @staticmethod
        def import_excel(tableid, file):
            result = ResultModel()
            result.msg = ''
            table = SysTableList.objects.get(id=tableid)
            if table == None:
                result.msg = '未找到指定表'
                return result
            if table.allowimport != 1:
                result.msg = '不允许导入'
                return result
            columns = SysTableColumn.objects.filter(Q(tableid=tableid) & Q(importvisible = 1))
            if(len(columns) < 1):
                result.msg = '无可导入的列'
                return result
            columnsdict = {}
            coldatatype = {}
            colprimarykey = {}
            colrequire = []
            for col in columns:
                columnsdict[col.description] = col.name
                coldatatype[col.name] = col.datatype
                colprimarykey[col.name] = col.primarkey
                if col.required:
                    colrequire.append(col.name)

            rootpath = ROOT_PATH
            dirname = time.strftime('%Y%m%d')
            datepath = 'upload\\temp\\' + dirname
            upload_dir = rootpath + '\\' + datepath
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir)
            filename = table.description + '_' + time.strftime('%Y%m%d%H%M%S') + '.xlsx'
            fileurl = '/upload/temp/' + dirname + '/' + filename
            filepath = upload_dir + '\\' + filename

            #保存上传的文件
            f = open(filepath, 'wb')
            for chunk in file.chunks():
                f.write(chunk)
            f.close()

            #加载excel文件
            wb = load_workbook(filepath)
            sheets = wb.get_sheet_names()
            sheet_first = sheets[0]
            ws = wb.get_sheet_by_name(sheet_first)
            table_rows = ws.rows
            table_columns = ws.columns
            values = []
            rowcount = 1
            table_head = []
            colcount = 0
            for row in table_rows:
                #第一行表头
                if rowcount == 1:
                    for c in row:
                        cn_colname = c.value
                        table_head.append(cn_colname)
                        if  not cn_colname or cn_colname == 'None' or cn_colname not in columnsdict.keys():
                            result.msg +=' 不存在 "' + str(cn_colname) + '" 列 </br>'
                            continue
                    differ = list(set(colrequire).difference(set(table_head)))

                    if len(differ) > 0:
                        result.msg +=' excel中必须包含 "' + ','.join(differ) + '" 列且有值 </br>'
                    if result.msg != '':
                        return result
                else:
                    value = {}
                    colcount = 0
                    for col in row:
                        cn_colname = table_head[colcount]
                        if not col.value :
                            result.msg += ' 第' + str(rowcount) + '行,' + cn_colname + '" 列值不允许为空 </br>'
                            continue

                        english_colname = columnsdict[cn_colname]

                        if coldatatype[english_colname] == 'out':
                            outvalue = SysHelper.get_out_value_id(tableid,english_colname,col.value)
                            if outvalue == '无':
                                result.msg+=' 第' + str(rowcount) + '行,' + cn_colname + ':' + col.value + ' 错误,未查询到值,'
                            else:
                                value[english_colname] = outvalue
                        elif colprimarykey[english_colname] == 1:
                            #当为插入模式时，要检测为主键的值不能已存在
                            if table.importtype == TableImportType.insert.value:
                                value_exist = SqlHelper.single('select count(*) from {0} where {1}=\'{2}\''.format(table.name,english_colname,col.value))
                                if value_exist != '0':
                                       result.msg+=' 第' + str(rowcount) + '行,' + cn_colname + ':' + col.value + ' 错误,该字段为主键，值已存在,'
                            #当为更新模式时，要检测为主键的值必须已存在 因为要当条件
                            elif table.importtype == TableImportType.update.value:
                                value_exist = SqlHelper.single('select count(*) from {0} where {1}=\'{2}\''.format(table.name,english_colname,col.value))
                                if value_exist == '0':
                                       result.msg+=' 第' + str(rowcount) + '行,' + cn_colname + ':' + col.value + ' 错误,该字段为主键，值不存在,'
                            value[english_colname] = col.value
                        else:
                            value[english_colname] = col.value
                        colcount = colcount + 1
                    values.append(value)

                rowcount = rowcount + 1
            #如果有错误信息 直接返回结果不继续执行
            if result.msg != '':
                    return result

            sqllist = []
            if  table.importtype == TableImportType.insert.value:
                sql = 'insert into {0}({1}) values({2})'
                for row in values:
                    addmodel = {}
                    for key,value in row.items():
                        addmodel[key] = str(value)

                    addmodel['CreateDateTime'] = time.strftime("%Y-%m-%d %H:%M:%S")
                    addmodel['ModifyDateTime'] = time.strftime("%Y-%m-%d %H:%M:%S")
                    addmodel['Creator'] = SysHelper.userid
                    addmodel['Modifier'] = SysHelper.userid
                    insert_values = ''
                    for v in addmodel.values():
                        insert_values+="'" + v + "',"
                    insert_values = insert_values.strip(',')
                    sql = sql.format(table.name,','.join(addmodel.keys()),insert_values)
                    sqllist.append(sql)
                result.flag = SqlHelper.bulk_execute(sqllist) == len(sqllist)
            elif table.importtype == TableImportType.update.value:
                sql = 'update {0} set {1} where {2}'
                primarykeys = SysHelper.get_column_names(tableid, "PrimarKey=1", "ListOrder")
                if len(primarykeys) <= 0:
                    result.msg = '请设置主键，因为导入类型为更新'
                    return result
                #更新条件
                for row in values:
                    newvalues = ''
                    condition = ''
                    for key,value in row.items():
                        if key in primarykeys:
                            condition = key + "= '" + str(value) + "' and "
                        newvalues += key + "='" + str(value) + "',"
                    newvalues +="ModifyDateTime='" + time.strftime("%Y-%m-%d %H:%M:%S") + "',"
                    newvalues +="Modifier='" + SysHelper.userid + "'"
                    condition+=" 1=1 "
                    sqllist.append(sql.format(table.name,newvalues,condition))
                result.flag = SqlHelper.bulk_execute(sqllist) >= len(sqllist)
            else:
                result.msg = '请设置表管理中导入类型'
            return result
        @staticmethod
        def export_excel(tableid):
            result = ResultModel()
            table = SysTableList.objects.get(id=tableid)
            if table == None:
                result.msg = '未找到指定表'
                return result
            if table.allowexport != 1:
                result.msg = '不允许导出'
                return result
            rootpath = sys.path[0] + "\\kingWeb"
            dirname = time.strftime('%Y%m%d')
            datepath = 'upload\\temp\\' + dirname
            upload_dir = rootpath + '\\' + datepath
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir)
            filename = table.description + '_' + time.strftime('%Y%m%d%H%M%S') + '.xlsx'
            fileurl = '/upload/temp/' + dirname + '/' + filename
            filepath = upload_dir + '\\' + filename

            wb = Workbook()
            ws = wb.active
            ws.title = table.description
            columns = SysTableColumn.objects.filter(Q(tableid=tableid) & Q(exportvisible = 1))
            columnstr = ''
            table_head = []   #表头
            coldatatype = {}
            for col in columns:
                table_head.append(col.description)
                columnstr+=col.name + ','
                coldatatype[col.name] = col.datatype
            columnstr = columnstr.rstrip(',')
            ws.append(table_head)
            tabledata = SqlHelper.query('select {0} from {1}'.format(columnstr,table.name))
            for row in tabledata:
                rowdata = []
                for key,value in row.items():
                    if coldatatype[key] == 'out':
                        rowdata.append(SysHelper.get_out_value(tableid,key,value))
                    else:
                        rowdata.append(str(value))
                ws.append(rowdata)

            wb.save(filename=filepath)
            result.data = fileurl
            result.flag = True
            return result
        @staticmethod
        def download_import_template(tableid):
            result = ResultModel()
            table = SysTableList.objects.get(id=tableid)
            if table == None:
                result.msg = '未找到指定表'
                return result
            if table.allowimport != 1:
                result.msg = '不允许导入'
                return result
            rootpath = ROOT_PATH
            dirname = time.strftime('%Y%m%d')
            datepath = 'upload\\temp\\' + dirname
            upload_dir = rootpath + '\\' + datepath
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir)
            filename = table.description + '_导入模板_' + time.strftime('%Y%m%d%H%M%S') + '.xlsx'
            fileurl = '/upload/temp/' + dirname + '/' + filename
            filepath = upload_dir + '\\' + filename

            wb = Workbook()
            ws = wb.active
            ws.title = table.description
            columns = SysTableColumn.objects.filter(Q(tableid=tableid) & Q(importvisible = 1))
            columnstr = ''
            table_head = []   #表头
            for col in columns:
                table_head.append(col.description)
            ws.append(table_head)

            wb.save(filename=filepath)
            result.data = fileurl
            result.flag = True
            return result